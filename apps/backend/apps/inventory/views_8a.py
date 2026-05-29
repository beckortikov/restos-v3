"""Phase 8A — viewsets для поставщиков, накладных, списаний, расхода,
инвентаризации + XLSX-импорт/экспорт.
"""
from __future__ import annotations

import io
from datetime import date as _date
from decimal import Decimal, InvalidOperation

from django.http import HttpResponse
from rest_framework import status as http_status
from rest_framework import viewsets
from rest_framework.decorators import action
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response

from common.exceptions import BusinessError
from common.permissions import IsCashier

from .models import (
    Ingredient,
    InventoryCheck,
    StockReceipt,
    StockWriteoff,
    Supplier,
    SupplyExpense,
)
from .serializers import (
    InventoryCheckSerializer,
    StockReceiptSerializer,
    StockWriteoffSerializer,
    SupplierSerializer,
    SupplyExpenseSerializer,
)
from .services_8a import (
    apply_inventory_check,
    apply_receipt,
    apply_writeoff,
    populate_inventory_check_from_stock,
    record_supply_expense,
)


# ─── Supplier ───────────────────────────────────────────────────────────────


class SupplierViewSet(viewsets.ModelViewSet):
    serializer_class = SupplierSerializer
    permission_classes = [IsCashier]

    def get_queryset(self):
        qs = Supplier.objects.filter(restaurant=self.request.user.restaurant)
        active = self.request.query_params.get("is_active")
        if active is not None:
            qs = qs.filter(is_active=active.lower() in ("1", "true", "yes"))
        return qs

    def perform_create(self, serializer):
        serializer.save(restaurant=self.request.user.restaurant)


# ─── Stock Receipts ─────────────────────────────────────────────────────────


class StockReceiptViewSet(viewsets.ModelViewSet):
    serializer_class = StockReceiptSerializer
    permission_classes = [IsCashier]

    def get_queryset(self):
        qs = StockReceipt.objects.filter(
            restaurant=self.request.user.restaurant
        ).select_related("supplier").prefetch_related("lines__ingredient")
        status_p = self.request.query_params.get("status")
        if status_p:
            qs = qs.filter(status=status_p)
        return qs.order_by("-receipt_date", "-id")

    def perform_create(self, serializer):
        serializer.save(restaurant=self.request.user.restaurant)

    @action(detail=True, methods=["post"], url_path="apply")
    def apply(self, request, pk=None):
        receipt = self.get_object()
        apply_receipt(receipt, user=request.user)
        return Response({"data": StockReceiptSerializer(receipt).data})


# ─── Writeoffs ──────────────────────────────────────────────────────────────


class StockWriteoffViewSet(viewsets.ModelViewSet):
    serializer_class = StockWriteoffSerializer
    permission_classes = [IsCashier]

    def get_queryset(self):
        qs = StockWriteoff.objects.filter(
            restaurant=self.request.user.restaurant
        ).prefetch_related("lines__ingredient")
        status_p = self.request.query_params.get("status")
        if status_p:
            qs = qs.filter(status=status_p)
        return qs.order_by("-writeoff_date", "-id")

    def perform_create(self, serializer):
        serializer.save(restaurant=self.request.user.restaurant)

    @action(detail=True, methods=["post"], url_path="apply")
    def apply(self, request, pk=None):
        wo = self.get_object()
        apply_writeoff(wo, user=request.user)
        return Response({"data": StockWriteoffSerializer(wo).data})


# ─── Supply Expense ─────────────────────────────────────────────────────────


class SupplyExpenseViewSet(viewsets.ReadOnlyModelViewSet):
    """Только READ + custom create через сервис (для проверок)."""

    serializer_class = SupplyExpenseSerializer
    permission_classes = [IsCashier]

    def get_queryset(self):
        return SupplyExpense.objects.filter(
            restaurant=self.request.user.restaurant
        ).select_related("ingredient", "user").order_by("-created_at")

    def create(self, request, *args, **kwargs):
        try:
            ing_id = int(request.data.get("ingredient"))
            qty = Decimal(str(request.data.get("qty") or 0))
        except (TypeError, ValueError, InvalidOperation):
            raise BusinessError("INVALID_VALUE", "ingredient и qty обязательны", 400)
        ingredient = Ingredient.objects.filter(
            id=ing_id, restaurant=request.user.restaurant,
        ).first()
        if ingredient is None:
            raise BusinessError("NOT_FOUND", "Ингредиент не найден", 404)
        reason = request.data.get("reason") or "household"
        note = (request.data.get("note") or "").strip()
        expense = record_supply_expense(
            restaurant=request.user.restaurant,
            ingredient=ingredient,
            qty=qty,
            reason=reason,
            note=note,
            user=request.user,
        )
        return Response(
            {"data": SupplyExpenseSerializer(expense).data},
            status=http_status.HTTP_201_CREATED,
        )


# ─── Inventory Check ────────────────────────────────────────────────────────


class InventoryCheckViewSet(viewsets.ModelViewSet):
    serializer_class = InventoryCheckSerializer
    permission_classes = [IsCashier]

    def get_queryset(self):
        return InventoryCheck.objects.filter(
            restaurant=self.request.user.restaurant
        ).prefetch_related("lines__ingredient").order_by("-check_date", "-id")

    def perform_create(self, serializer):
        check = serializer.save(restaurant=self.request.user.restaurant)
        # При создании сразу заполняем строки текущим остатком.
        populate_inventory_check_from_stock(check)

    @action(detail=True, methods=["post"], url_path="apply")
    def apply(self, request, pk=None):
        check = self.get_object()
        apply_inventory_check(check, user=request.user)
        return Response({"data": InventoryCheckSerializer(check).data})

    @action(detail=True, methods=["patch"], url_path="lines")
    def update_lines(self, request, pk=None):
        """Массово обновить actual_qty по линиям. body = [{id, actual_qty}, ...]."""
        check = self.get_object()
        if check.status != "draft":
            raise BusinessError("INVALID_STATE", "Документ проведён", 400)
        updates = request.data if isinstance(request.data, list) else (
            request.data.get("lines") or []
        )
        by_id = {int(u["id"]): u for u in updates if u.get("id") is not None}
        if not by_id:
            return Response({"data": InventoryCheckSerializer(check).data})
        from django.db import transaction

        with transaction.atomic():
            for line in check.lines.all():
                u = by_id.get(line.id)
                if u is None:
                    continue
                try:
                    line.actual_qty = Decimal(str(u.get("actual_qty", line.actual_qty)))
                    line.save(update_fields=["actual_qty"])
                except (TypeError, InvalidOperation):
                    raise BusinessError(
                        "INVALID_VALUE",
                        f"actual_qty некорректен для строки #{line.id}", 400,
                    )
        return Response({"data": InventoryCheckSerializer(check).data})


# ─── XLSX import/export ─────────────────────────────────────────────────────


def _xlsx_response(workbook, filename: str) -> HttpResponse:
    buf = io.BytesIO()
    workbook.save(buf)
    buf.seek(0)
    resp = HttpResponse(
        buf.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp["Content-Disposition"] = f'attachment; filename="{filename}"'
    return resp


@action(detail=False, methods=["get"], url_path="template")
def _ingredients_template_stub():  # placeholder for IDE
    pass


class IngredientsImportMixin:
    """Подмешивается в IngredientViewSet — даёт endpoints для шаблона и импорта."""

    @action(detail=False, methods=["get"], url_path="template", permission_classes=[IsCashier])
    def template(self, request):
        """GET /inventory/ingredients/template/ → XLSX-шаблон для импорта."""
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill

        wb = Workbook()
        ws = wb.active
        ws.title = "Ингредиенты"
        headers = [
            "Название*", "Единица*", "Низкий остаток", "Активен (1/0)", "Тип (food/household)",
        ]
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="F47A20")
        # Примеры
        ws.append(["Говядина", "kg", 2, 1, "food"])
        ws.append(["Туалетная бумага", "piece", 10, 1, "household"])
        ws.append(["Соль", "g", 500, 1, "food"])
        # Лист «Единицы» — справочно
        ws_units = wb.create_sheet("Единицы")
        ws_units.append(["Код", "Описание"])
        for code, label in [
            ("kg", "Килограмм"), ("g", "Грамм"),
            ("l", "Литр"), ("ml", "Миллилитр"),
            ("piece", "Штука"), ("pack", "Упаковка"), ("bottle", "Бутылка"),
        ]:
            ws_units.append([code, label])
        ws.column_dimensions["A"].width = 28
        for col in "BCDE":
            ws.column_dimensions[col].width = 18
        return _xlsx_response(wb, "ingredients_template.xlsx")

    @action(detail=False, methods=["post"], url_path="import", permission_classes=[IsCashier])
    def import_xlsx(self, request):
        """POST /inventory/ingredients/import/ multipart file=<xlsx>.

        Создаёт новые ингредиенты и обновляет существующие по полю name.
        Не трогает остатки (только справочник). Возвращает summary.
        """
        from openpyxl import load_workbook

        f = request.FILES.get("file")
        if f is None:
            raise BusinessError("INVALID_VALUE", "Файл не передан (file=)", 400)
        try:
            wb = load_workbook(filename=f, read_only=True, data_only=True)
        except Exception as e:
            raise BusinessError("INVALID_FILE", f"Не XLSX: {e}", 400)
        ws = wb.active

        VALID_UNITS = {"kg", "g", "l", "ml", "piece", "pack", "bottle"}
        created = 0
        updated = 0
        errors: list[dict] = []
        rid = request.user.restaurant_id
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        for idx, row in enumerate(rows, start=2):
            if not row or not any(row):
                continue
            try:
                name = (row[0] or "").strip() if row[0] else ""
                unit_raw = (row[1] or "").strip().lower() if row[1] else ""
                threshold = row[2] if len(row) > 2 else None
                active_raw = row[3] if len(row) > 3 else 1
                kind_raw = (row[4] or "food").strip().lower() if len(row) > 4 and row[4] else "food"
            except Exception as e:
                errors.append({"row": idx, "error": f"парсинг: {e}"})
                continue
            if not name:
                errors.append({"row": idx, "error": "пустое название"})
                continue
            if unit_raw not in VALID_UNITS:
                errors.append({
                    "row": idx, "error": f"неизвестная единица '{unit_raw}'",
                })
                continue
            try:
                thr_value = (
                    Decimal(str(threshold)) if threshold not in (None, "", "—") else None
                )
            except (TypeError, InvalidOperation):
                thr_value = None
            is_active = bool(active_raw) and str(active_raw).strip() not in ("0", "false", "False", "no")
            is_food = kind_raw not in ("household", "supply", "non_food", "hozhtovar", "хозтовар")

            existing = Ingredient.objects.filter(
                restaurant_id=rid, name=name,
            ).first()
            if existing:
                existing.unit = unit_raw
                existing.low_stock_threshold = thr_value
                existing.is_active = is_active
                existing.is_food = is_food
                existing.save(update_fields=[
                    "unit", "low_stock_threshold", "is_active", "is_food",
                    "updated_at",
                ])
                updated += 1
            else:
                Ingredient.objects.create(
                    restaurant_id=rid, name=name, unit=unit_raw,
                    low_stock_threshold=thr_value,
                    is_active=is_active, is_food=is_food,
                )
                created += 1

        return Response({
            "data": {
                "created": created, "updated": updated,
                "errors": errors, "total_rows": len(rows),
            }
        })


class ReceiptsImportMixin:
    """Mixin для StockReceiptViewSet — шаблон и импорт накладной из XLSX."""

    @action(detail=False, methods=["get"], url_path="template", permission_classes=[IsCashier])
    def template(self, request):
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill

        wb = Workbook()
        ws = wb.active
        ws.title = "Позиции"
        headers = ["Ингредиент* (имя)", "Кол-во*", "Цена за единицу*"]
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="F47A20")
        ws.append(["Говядина", 10, 100])
        ws.append(["Лук", 5, 12])
        ws.append(["Соль", 1000, 0.05])
        for col, w in [("A", 30), ("B", 14), ("C", 18)]:
            ws.column_dimensions[col].width = w

        # Лист с метаданными накладной — необязательный, но позволяет
        # импортировать накладную одним файлом.
        ws_meta = wb.create_sheet("Накладная")
        ws_meta.append(["Поле", "Значение"])
        for cell in ws_meta[1]:
            cell.font = Font(bold=True)
        ws_meta.append(["Поставщик (имя)", "ИП Бахром"])
        ws_meta.append(["Дата (YYYY-MM-DD)", _date.today().isoformat()])
        ws_meta.append(["Номер документа", "001"])
        ws_meta.append(["Заметка", ""])
        ws_meta.column_dimensions["A"].width = 28
        ws_meta.column_dimensions["B"].width = 30
        return _xlsx_response(wb, "receipt_template.xlsx")

    @action(detail=False, methods=["post"], url_path="import", permission_classes=[IsCashier])
    def import_xlsx(self, request):
        """POST /inventory/receipts/import/ — создаёт DRAFT-накладную из XLSX.

        Лист «Позиции»: имя, qty, unit_cost.
        Лист «Накладная» (опц.): поставщик/дата/номер.
        """
        from openpyxl import load_workbook

        from .models import StockReceipt, StockReceiptLine, Supplier

        f = request.FILES.get("file")
        if f is None:
            raise BusinessError("INVALID_VALUE", "Файл не передан (file=)", 400)
        try:
            wb = load_workbook(filename=f, read_only=True, data_only=True)
        except Exception as e:
            raise BusinessError("INVALID_FILE", f"Не XLSX: {e}", 400)

        # Метаданные накладной
        supplier_name = ""
        receipt_date = _date.today()
        number = ""
        note = ""
        if "Накладная" in wb.sheetnames:
            ws_meta = wb["Накладная"]
            for row in ws_meta.iter_rows(min_row=2, values_only=True):
                if not row or not row[0]:
                    continue
                key = (row[0] or "").strip().lower()
                val = (row[1] or "").strip() if row[1] else ""
                if "поставщик" in key:
                    supplier_name = val
                elif "дата" in key:
                    try:
                        receipt_date = _date.fromisoformat(val)
                    except (TypeError, ValueError):
                        pass
                elif "номер" in key:
                    number = val
                elif "заметка" in key or "примечание" in key:
                    note = val

        rid = request.user.restaurant_id
        supplier = None
        if supplier_name:
            supplier, _ = Supplier.objects.get_or_create(
                restaurant_id=rid, name=supplier_name,
                defaults={"is_active": True},
            )

        # Позиции
        ws_lines = wb["Позиции"] if "Позиции" in wb.sheetnames else wb.active
        parsed: list[dict] = []
        errors: list[dict] = []
        for idx, row in enumerate(ws_lines.iter_rows(min_row=2, values_only=True), start=2):
            if not row or not any(row):
                continue
            ing_name = (row[0] or "").strip() if row[0] else ""
            if not ing_name:
                continue
            try:
                qty = Decimal(str(row[1]))
                cost = Decimal(str(row[2]))
            except (TypeError, InvalidOperation):
                errors.append({"row": idx, "error": "qty/unit_cost не число"})
                continue
            ing = Ingredient.objects.filter(restaurant_id=rid, name=ing_name).first()
            if ing is None:
                errors.append({"row": idx, "error": f"ингредиент '{ing_name}' не найден"})
                continue
            parsed.append({"ingredient": ing, "qty": qty, "unit_cost": cost})

        if not parsed:
            raise BusinessError(
                "INVALID_FILE",
                f"Нет валидных позиций. Ошибки: {errors[:5]}", 400,
            )

        from django.db import transaction

        with transaction.atomic():
            receipt = StockReceipt.objects.create(
                restaurant_id=rid, supplier=supplier,
                receipt_date=receipt_date, number=number, note=note,
                status="draft", created_by=request.user,
            )
            total = Decimal("0")
            for ln in parsed:
                line_total = (ln["qty"] * ln["unit_cost"]).quantize(Decimal("0.01"))
                StockReceiptLine.objects.create(
                    receipt=receipt,
                    ingredient=ln["ingredient"],
                    qty=ln["qty"],
                    unit_cost=ln["unit_cost"],
                    total=line_total,
                )
                total += line_total
            receipt.total_amount = total.quantize(Decimal("0.01"))
            receipt.save(update_fields=["total_amount"])

        return Response({
            "data": StockReceiptSerializer(receipt).data,
            "meta": {"errors": errors, "imported": len(parsed)},
        }, status=http_status.HTTP_201_CREATED)
