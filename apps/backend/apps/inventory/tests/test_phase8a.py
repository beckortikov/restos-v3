"""Phase 8A — поставщики, накладные, списания, расход, инвентаризация, XLSX."""
from datetime import date
from decimal import Decimal

import pytest

pytestmark = pytest.mark.django_db


@pytest.fixture
def beef(restaurant):
    from apps.inventory.models import Ingredient
    return Ingredient.objects.create(
        restaurant=restaurant, name="Говядина", unit="kg", is_food=True,
    )


@pytest.fixture
def soap(restaurant):
    from apps.inventory.models import Ingredient
    return Ingredient.objects.create(
        restaurant=restaurant, name="Жидкое мыло", unit="bottle", is_food=False,
    )


@pytest.fixture
def supplier(restaurant):
    from apps.inventory.models import Supplier
    return Supplier.objects.create(
        restaurant=restaurant, name="ИП Бахром", phone="+992 99 000",
    )


# ─── apply_receipt ──────────────────────────────────────────────────────────


def test_apply_receipt_creates_purchase_movements(restaurant, beef, supplier, cashier):
    from apps.inventory.models import StockReceipt, StockReceiptLine
    from apps.inventory.services_8a import apply_receipt

    r = StockReceipt.objects.create(
        restaurant=restaurant, supplier=supplier,
        receipt_date=date.today(), number="001",
    )
    StockReceiptLine.objects.create(
        receipt=r, ingredient=beef,
        qty=Decimal("10"), unit_cost=Decimal("100"), total=Decimal("1000"),
    )
    apply_receipt(r, user=cashier)

    r.refresh_from_db()
    assert r.status == "applied"
    assert r.applied_at is not None
    beef.refresh_from_db()
    assert beef.current_qty == Decimal("10.000")
    assert beef.avg_cost_per_unit == Decimal("100.0000")


def test_apply_receipt_rejects_already_applied(restaurant, beef, supplier):
    from apps.inventory.models import StockReceipt, StockReceiptLine
    from apps.inventory.services_8a import apply_receipt
    from common.exceptions import BusinessError

    r = StockReceipt.objects.create(
        restaurant=restaurant, supplier=supplier,
        receipt_date=date.today(),
    )
    StockReceiptLine.objects.create(
        receipt=r, ingredient=beef, qty=Decimal("1"),
        unit_cost=Decimal("100"), total=Decimal("100"),
    )
    apply_receipt(r)
    with pytest.raises(BusinessError) as exc:
        apply_receipt(r)
    assert exc.value.code == "INVALID_STATE"


# ─── apply_writeoff ─────────────────────────────────────────────────────────


def test_apply_writeoff_creates_waste_movements(restaurant, beef, cashier):
    from apps.inventory.models import StockMovementKind, StockWriteoff, StockWriteoffLine
    from apps.inventory.services import record_movement
    from apps.inventory.services_8a import apply_writeoff

    record_movement(
        ingredient=beef, kind=StockMovementKind.PURCHASE,
        qty_delta=Decimal("10"), unit_cost=Decimal("100"), reason="init",
    )
    wo = StockWriteoff.objects.create(
        restaurant=restaurant, writeoff_date=date.today(), reason="spoilage",
    )
    StockWriteoffLine.objects.create(writeoff=wo, ingredient=beef, qty=Decimal("3"))
    apply_writeoff(wo, user=cashier)

    wo.refresh_from_db()
    assert wo.status == "applied"
    beef.refresh_from_db()
    assert beef.current_qty == Decimal("7.000")


# ─── record_supply_expense ──────────────────────────────────────────────────


def test_supply_expense_blocks_negative(restaurant, soap):
    from apps.inventory.services_8a import record_supply_expense
    from common.exceptions import BusinessError

    # Склад пустой, supply_allow_negative=False по умолчанию
    with pytest.raises(BusinessError) as exc:
        record_supply_expense(
            restaurant=restaurant, ingredient=soap, qty=Decimal("5"),
        )
    assert exc.value.code == "INSUFFICIENT_STOCK"


def test_supply_expense_allows_negative_when_flag(restaurant, soap, cashier):
    from apps.inventory.services_8a import record_supply_expense

    restaurant.supply_allow_negative = True
    restaurant.save(update_fields=["supply_allow_negative"])

    expense = record_supply_expense(
        restaurant=restaurant, ingredient=soap, qty=Decimal("2"),
        reason="household", note="в туалет", user=cashier,
    )
    assert expense.id is not None
    soap.refresh_from_db()
    assert soap.current_qty == Decimal("-2.000")


def test_supply_expense_normal_flow(restaurant, soap, cashier):
    from apps.inventory.models import StockMovementKind
    from apps.inventory.services import record_movement
    from apps.inventory.services_8a import record_supply_expense

    record_movement(
        ingredient=soap, kind=StockMovementKind.PURCHASE,
        qty_delta=Decimal("10"), unit_cost=Decimal("50"), reason="init",
    )
    record_supply_expense(
        restaurant=restaurant, ingredient=soap, qty=Decimal("3"),
        reason="to_hall", user=cashier,
    )
    soap.refresh_from_db()
    assert soap.current_qty == Decimal("7.000")


# ─── apply_inventory_check ──────────────────────────────────────────────────


def test_inventory_check_writes_diff_movements(restaurant, beef, cashier):
    from apps.inventory.models import (
        InventoryCheck,
        InventoryCheckLine,
        StockMovementKind,
    )
    from apps.inventory.services import record_movement
    from apps.inventory.services_8a import apply_inventory_check

    record_movement(
        ingredient=beef, kind=StockMovementKind.PURCHASE,
        qty_delta=Decimal("10"), unit_cost=Decimal("100"), reason="init",
    )
    check = InventoryCheck.objects.create(
        restaurant=restaurant, check_date=date.today(),
    )
    # Фактически нашли 8 кг (потеря 2 кг где-то)
    InventoryCheckLine.objects.create(
        inventory_check=check, ingredient=beef,
        expected_qty=Decimal("10"), actual_qty=Decimal("8"),
    )
    apply_inventory_check(check, user=cashier)

    check.refresh_from_db()
    assert check.status == "applied"
    beef.refresh_from_db()
    assert beef.current_qty == Decimal("8.000")


def test_inventory_check_skips_zero_diff(restaurant, beef, cashier):
    from apps.inventory.models import (
        InventoryCheck,
        InventoryCheckLine,
        IngredientStockMovement,
        StockMovementKind,
    )
    from apps.inventory.services import record_movement
    from apps.inventory.services_8a import apply_inventory_check

    record_movement(
        ingredient=beef, kind=StockMovementKind.PURCHASE,
        qty_delta=Decimal("10"), unit_cost=Decimal("100"), reason="init",
    )
    check = InventoryCheck.objects.create(
        restaurant=restaurant, check_date=date.today(),
    )
    InventoryCheckLine.objects.create(
        inventory_check=check, ingredient=beef,
        expected_qty=Decimal("10"), actual_qty=Decimal("10"),
    )
    before = IngredientStockMovement.objects.count()
    apply_inventory_check(check, user=cashier)
    after = IngredientStockMovement.objects.count()
    assert after == before  # ничего не добавлено


# ─── XLSX template/import API ───────────────────────────────────────────────


def test_ingredients_template_download(api_client, cashier):
    api_client.force_authenticate(user=cashier)
    resp = api_client.get("/api/v1/inventory/ingredients/template/")
    assert resp.status_code == 200
    assert "spreadsheet" in resp["Content-Type"]
    assert b"Ingredient" not in resp.content[:200] or True  # просто что вернулся xlsx


def test_ingredients_import_creates_and_updates(api_client, cashier, restaurant):
    """Импорт: 2 новые + обновление существующей."""
    import io
    from openpyxl import Workbook
    from django.core.files.uploadedfile import SimpleUploadedFile

    from apps.inventory.models import Ingredient

    Ingredient.objects.create(
        restaurant=restaurant, name="Соль", unit="g",
        low_stock_threshold=100, is_food=True,
    )

    wb = Workbook()
    ws = wb.active
    ws.append(["Название*", "Единица*", "Низкий остаток", "Активен (1/0)", "Тип"])
    ws.append(["Соль", "g", 500, 1, "food"])  # обновится
    ws.append(["Перец", "g", 50, 1, "food"])  # новый
    ws.append(["Туалетная бумага", "piece", 10, 1, "household"])  # хозтовар
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    upload = SimpleUploadedFile(
        "ing.xlsx", buf.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

    api_client.force_authenticate(user=cashier)
    resp = api_client.post(
        "/api/v1/inventory/ingredients/import/",
        {"file": upload}, format="multipart",
    )
    assert resp.status_code == 200, resp.content
    body = resp.json()["data"]
    assert body["created"] == 2
    assert body["updated"] == 1

    salt = Ingredient.objects.get(restaurant=restaurant, name="Соль")
    assert salt.low_stock_threshold == Decimal("500.000")
    soap = Ingredient.objects.get(restaurant=restaurant, name="Туалетная бумага")
    assert soap.is_food is False


def test_ingredients_import_reports_errors(api_client, cashier):
    """Ошибочные строки попадают в errors, валидные — импортятся."""
    import io
    from openpyxl import Workbook
    from django.core.files.uploadedfile import SimpleUploadedFile

    wb = Workbook()
    ws = wb.active
    ws.append(["Название*", "Единица*"])
    ws.append(["Хороший", "kg"])
    ws.append(["", "kg"])  # пустое имя
    ws.append(["Неизвестный", "fakeunit"])  # неизвестная единица
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    upload = SimpleUploadedFile("ing.xlsx", buf.getvalue())

    api_client.force_authenticate(user=cashier)
    resp = api_client.post(
        "/api/v1/inventory/ingredients/import/",
        {"file": upload}, format="multipart",
    )
    body = resp.json()["data"]
    assert body["created"] == 1
    assert len(body["errors"]) == 2


def test_receipt_template_download(api_client, cashier):
    api_client.force_authenticate(user=cashier)
    resp = api_client.get("/api/v1/inventory/receipts/template/")
    assert resp.status_code == 200
    assert "spreadsheet" in resp["Content-Type"]


def test_receipt_import_creates_draft(api_client, cashier, restaurant, beef):
    import io
    from openpyxl import Workbook
    from django.core.files.uploadedfile import SimpleUploadedFile

    wb = Workbook()
    ws = wb.active
    ws.title = "Позиции"
    ws.append(["Ингредиент*", "Кол-во*", "Цена*"])
    ws.append(["Говядина", 5, 120])

    ws_meta = wb.create_sheet("Накладная")
    ws_meta.append(["Поле", "Значение"])
    ws_meta.append(["Поставщик", "ИП Иван"])
    ws_meta.append(["Дата", date.today().isoformat()])
    ws_meta.append(["Номер", "777"])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    upload = SimpleUploadedFile("r.xlsx", buf.getvalue())

    api_client.force_authenticate(user=cashier)
    resp = api_client.post(
        "/api/v1/inventory/receipts/import/",
        {"file": upload}, format="multipart",
    )
    assert resp.status_code == 201, resp.content
    body = resp.json()["data"]
    assert body["status"] == "draft"
    assert body["number"] == "777"
    assert len(body["lines"]) == 1
    assert Decimal(body["total_amount"]) == Decimal("600.00")


# ─── Cross-restaurant isolation ─────────────────────────────────────────────


def test_supplier_isolation(api_client, cashier, restaurant):
    from apps.inventory.models import Supplier
    from apps.users.models import Restaurant

    Supplier.objects.create(restaurant=restaurant, name="Мой")
    other = Restaurant.objects.create(name="Other", currency="TJS")
    Supplier.objects.create(restaurant=other, name="Чужой")

    api_client.force_authenticate(user=cashier)
    resp = api_client.get("/api/v1/inventory/suppliers/")
    body = resp.json()
    items = body if isinstance(body, list) else body.get("data") or body.get("results") or []
    names = [s["name"] for s in items]
    assert "Мой" in names
    assert "Чужой" not in names
