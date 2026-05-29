from rest_framework import mixins, status, viewsets
from rest_framework.decorators import action
from rest_framework.response import Response

from common.permissions import IsCashier, IsCashierOrWaiter

from .models import Category, MenuItem, MenuItemNote, ModifierGroup
from .serializers import (
    CategorySerializer,
    MenuItemNoteSerializer,
    MenuItemSerializer,
    ModifierGroupSerializer,
)


class CategoryViewSet(
    mixins.ListModelMixin,
    mixins.CreateModelMixin,
    mixins.UpdateModelMixin,
    mixins.DestroyModelMixin,
    viewsets.GenericViewSet,
):
    """Категории меню. CRUD доступен кассиру (для frame 19. Настройки — Меню)."""

    serializer_class = CategorySerializer
    pagination_class = None

    def get_queryset(self):
        return Category.objects.filter(restaurant=self.request.user.restaurant)

    def get_permissions(self):
        if self.action in {"create", "update", "partial_update", "destroy"}:
            return [IsCashier()]
        return [IsCashierOrWaiter()]

    def list(self, request, *args, **kwargs):
        qs = self.filter_queryset(self.get_queryset())
        return Response(
            {"data": CategorySerializer(qs, many=True).data, "meta": {"total": qs.count()}}
        )

    def perform_create(self, serializer):
        serializer.save(restaurant=self.request.user.restaurant)


class MenuItemViewSet(
    mixins.ListModelMixin,
    mixins.RetrieveModelMixin,
    mixins.CreateModelMixin,
    mixins.UpdateModelMixin,
    mixins.DestroyModelMixin,
    viewsets.GenericViewSet,
):
    """Блюда. Toggle is_available (стоп-лист) — PATCH `/menu/items/{id}/`."""

    serializer_class = MenuItemSerializer
    filterset_fields = ["category", "is_available"]
    pagination_class = None

    def get_queryset(self):
        return MenuItem.objects.filter(
            restaurant=self.request.user.restaurant
        ).select_related("category")

    def get_permissions(self):
        if self.action in {
            "create", "update", "partial_update", "destroy",
            "toggle_available", "stop_list", "restore",
            "allow_oversell", "toggle_tech_card",
            "auto_stopped_list",
        }:
            return [IsCashier()]
        return [IsCashierOrWaiter()]

    def list(self, request, *args, **kwargs):
        qs = self.filter_queryset(self.get_queryset())
        last = qs.order_by("-updated_at").values_list("updated_at", flat=True).first()
        cnt = qs.count()
        ts = f"{last.timestamp():.6f}" if last else "0"
        etag = f'W/"{ts}-{cnt}"'

        if request.headers.get("If-None-Match") == etag:
            resp = Response(status=304)
        else:
            data = MenuItemSerializer(qs, many=True, context={"request": request}).data
            resp = Response({"data": data, "meta": {"total": cnt}})

        resp["ETag"] = etag
        resp["Cache-Control"] = "max-age=300"
        return resp

    def retrieve(self, request, *args, **kwargs):
        item = self.get_object()
        return Response(
            {"data": MenuItemSerializer(item, context={"request": request}).data}
        )

    def perform_create(self, serializer):
        serializer.save(restaurant=self.request.user.restaurant)

    @action(detail=True, methods=["post"], url_path="toggle_available")
    def toggle_available(self, request, pk=None):
        """Быстрый toggle is_available для frame 14 «Стоп-лист»."""
        item = self.get_object()
        item.is_available = not item.is_available
        if item.is_available:
            # При возврате в продажу очищаем поля стоп-листа
            item.stop_reason = ""
            item.stop_until = None
        item.save(
            update_fields=[
                "is_available", "stop_reason", "stop_until", "updated_at",
            ]
        )
        return Response(
            {"data": MenuItemSerializer(item, context={"request": request}).data}
        )

    @action(detail=True, methods=["post"], url_path="stop_list")
    def stop_list(self, request, pk=None):
        """Снять блюдо со стоп-листа: is_available=False + reason + until.

        Body: {reason: str, until: "YYYY-MM-DD" | null}.
        """
        from apps.audit.services import log_request

        item = self.get_object()
        reason = (request.data.get("reason") or "").strip()
        until_raw = request.data.get("until")
        until = None
        if until_raw:
            from datetime import date

            try:
                until = date.fromisoformat(str(until_raw))
            except (ValueError, TypeError):
                from common.exceptions import BusinessError

                raise BusinessError(
                    "INVALID_TRANSITION",
                    "Дата возврата должна быть в формате YYYY-MM-DD", 422,
                )
        item.is_available = False
        item.stop_reason = reason
        item.stop_until = until
        item.save(
            update_fields=[
                "is_available", "stop_reason", "stop_until", "updated_at",
            ]
        )
        log_request(
            request, "settings_update", target=item,
            payload={
                "action": "stop_list",
                "menu_item": item.name,
                "reason": reason,
                "until": until.isoformat() if until else None,
            },
        )
        return Response(
            {"data": MenuItemSerializer(item, context={"request": request}).data}
        )

    @action(
        detail=False, methods=["get"], url_path="full",
        permission_classes=[IsCashierOrWaiter],
    )
    def full(self, request):
        """Категории + все доступные блюда одним запросом — для waiter PWA.

        Возвращает: {data: {categories: [...], items: [...]}, meta: {...}}.
        Только активные/доступные позиции (`is_available=True`). ETag-кэш
        по обоим таблицам — клиент кэширует целое меню на 5 минут.
        """
        from .models import Category
        from .serializers import CategorySerializer

        cats_qs = Category.objects.filter(
            restaurant=request.user.restaurant
        ).order_by("sort_order", "name")
        items_qs = MenuItem.objects.filter(
            restaurant=request.user.restaurant, is_available=True,
        ).select_related("category").order_by(
            "category__sort_order", "sort_order", "name",
        )

        last_cat = cats_qs.order_by(
            "-updated_at"
        ).values_list("updated_at", flat=True).first() if hasattr(
            Category, "updated_at"
        ) else None
        last_item = items_qs.order_by("-updated_at").values_list(
            "updated_at", flat=True,
        ).first()
        items_count = items_qs.count()
        cats_count = cats_qs.count()
        timestamps = [t.timestamp() for t in (last_cat, last_item) if t]
        last_ts = max(timestamps) if timestamps else 0
        etag = f'W/"full-{last_ts:.6f}-{cats_count}-{items_count}"'

        if request.headers.get("If-None-Match") == etag:
            resp = Response(status=304)
        else:
            resp = Response({
                "data": {
                    "categories": CategorySerializer(
                        cats_qs, many=True, context={"request": request},
                    ).data,
                    "items": MenuItemSerializer(
                        items_qs, many=True, context={"request": request},
                    ).data,
                },
                "meta": {
                    "categories_total": cats_count,
                    "items_total": items_count,
                },
            })
        resp["ETag"] = etag
        resp["Cache-Control"] = "max-age=300"
        return resp

    @action(
        detail=False, methods=["get"], url_path="template",
        permission_classes=[IsCashier],
    )
    def template_xlsx(self, request):
        """GET /menu/items/template/ → XLSX-шаблон для импорта меню."""
        import io
        from django.http import HttpResponse
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill

        wb = Workbook()
        ws = wb.active
        ws.title = "Menu"
        headers = ["Категория", "Блюдо", "Цена", "Эмодзи", "Доступно"]
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="F47A20")
        ws.append(["Горячее", "Плов", 45.00, "🍚", "да"])
        ws.append(["Горячее", "Лагман", 40.00, "🍜", "да"])
        ws.append(["Напитки", "Чай зелёный", 8.00, "🍵", "да"])
        ws.append(["Напитки", "Кола 0.5", 12.00, "🥤", "нет"])
        for col, w in [("A", 24), ("B", 30), ("C", 10), ("D", 10), ("E", 12)]:
            ws.column_dimensions[col].width = w

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        resp = HttpResponse(
            buf.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument."
                         "spreadsheetml.sheet",
        )
        resp["Content-Disposition"] = 'attachment; filename="menu_template.xlsx"'
        return resp

    @action(
        detail=False, methods=["post"], url_path="import_xlsx",
        permission_classes=[IsCashier],
    )
    def import_xlsx(self, request):
        """POST /menu/items/import_xlsx/  multipart с файлом `file`.

        Возвращает summary: {created, updated, errors: [...]}.
        """
        from apps.menu.import_excel import import_menu_xlsx
        from common.exceptions import BusinessError

        f = request.FILES.get("file")
        if f is None:
            raise BusinessError("FILE_REQUIRED", "Файл не передан", 400)
        try:
            data = f.read()
            summary = import_menu_xlsx(
                data,
                restaurant=request.user.restaurant,
                user=request.user,
            )
        except ValueError as e:
            raise BusinessError("INVALID_FILE", str(e), 422)
        except Exception as e:
            raise BusinessError(
                "IMPORT_FAILED", f"Не удалось импортировать: {e}", 500,
            )
        return Response({"data": summary})

    @action(detail=True, methods=["get", "put"], url_path="tech_card")
    def tech_card(self, request, pk=None):
        """GET/PUT техкарта блюда.

        GET — список строк техкарты.
        PUT — full upsert: { lines: [{ingredient|nested_semi, qty_per_unit, sort_order}] }.
        После PUT signal автоматически пересчитает MenuItem.cogs.
        """
        from apps.inventory.models import Ingredient, SemiFinishedType

        from .models import MenuItemTechCardLine
        from .serializers import MenuItemTechCardLineSerializer

        item = self.get_object()
        if request.method == "GET":
            qs = item.tech_card_lines.all().select_related(
                "ingredient", "nested_semi",
            )
            return Response({
                "data": MenuItemTechCardLineSerializer(qs, many=True).data,
                "meta": {"cogs": str(item.cogs)},
            })

        # PUT — пересоздаём все строки
        from common.exceptions import BusinessError

        lines = request.data.get("lines") or []
        if not isinstance(lines, list):
            raise BusinessError("INVALID_VALUE", "lines должен быть list", 400)

        rid = item.restaurant_id
        # Pre-validate перед deletion (атомарность)
        from decimal import Decimal as _D
        validated: list[dict] = []
        for ln in lines:
            ing_id = ln.get("ingredient")
            sem_id = ln.get("nested_semi")
            if (ing_id is None) == (sem_id is None):
                raise BusinessError(
                    "INVALID_VALUE",
                    "Ровно один компонент на строку: ingredient ИЛИ nested_semi",
                    400,
                )
            ing = sem = None
            if ing_id:
                ing = Ingredient.objects.filter(id=ing_id, restaurant_id=rid).first()
                if ing is None:
                    raise BusinessError(
                        "INVALID_VALUE",
                        f"Ingredient #{ing_id} не найден в вашем ресторане", 400,
                    )
            if sem_id:
                sem = SemiFinishedType.objects.filter(id=sem_id, restaurant_id=rid).first()
                if sem is None:
                    raise BusinessError(
                        "INVALID_VALUE",
                        f"Semi #{sem_id} не найден в вашем ресторане", 400,
                    )
            try:
                qpu = _D(str(ln.get("qty_per_unit", "0")))
            except Exception:
                raise BusinessError("INVALID_VALUE", "qty_per_unit не число", 400)
            if qpu <= 0:
                raise BusinessError("INVALID_VALUE", "qty_per_unit должно быть > 0", 400)
            validated.append({
                "ingredient": ing,
                "nested_semi": sem,
                "qty_per_unit": qpu,
                "sort_order": int(ln.get("sort_order", 0)),
            })

        from django.db import transaction

        with transaction.atomic():
            item.tech_card_lines.all().delete()
            for v in validated:
                MenuItemTechCardLine.objects.create(menu_item=item, **v)
        # signal автоматически пересчитает cogs

        item.refresh_from_db()
        qs = item.tech_card_lines.all().select_related("ingredient", "nested_semi")
        return Response({
            "data": MenuItemTechCardLineSerializer(qs, many=True).data,
            "meta": {"cogs": str(item.cogs)},
        })

    @action(detail=True, methods=["post", "get"], url_path="batch_cook")
    def batch_cook(self, request, pk=None):
        """Phase 7E: cook нажал «+N порций» / просмотр истории.

        POST { qty: int > 0, note?: str } → +N к prepared_qty, списать сырьё
        по техкарте × N (если есть), запись в BatchCookingLog (kind=COOK).

        GET  → список последних 50 записей BatchCookingLog по этому блюду.
        """
        from .models import BatchCookingLog
        from .serializers import BatchCookingLogSerializer
        from .services import record_batch_cook

        item = self.get_object()

        if request.method == "GET":
            qs = (
                BatchCookingLog.objects
                .filter(menu_item=item)
                .select_related("user")
                .order_by("-created_at")[:50]
            )
            return Response({
                "data": BatchCookingLogSerializer(qs, many=True).data,
                "meta": {"prepared_qty": item.prepared_qty},
            })

        qty_raw = request.data.get("qty")
        note = (request.data.get("note") or "").strip()
        try:
            qty = int(qty_raw)
        except (TypeError, ValueError):
            from common.exceptions import BusinessError
            raise BusinessError("INVALID_VALUE", "qty должно быть целым > 0", 400)
        if qty <= 0:
            from common.exceptions import BusinessError
            raise BusinessError("INVALID_VALUE", "qty должно быть > 0", 400)

        result = record_batch_cook(
            item,
            qty_delta=qty,
            kind="cook",
            user=request.user,
            note=note,
        )

        from apps.audit.services import log_request
        log_request(
            request, "batch_cook", target=item,
            payload={"qty": qty, "note": note, "new_total": result["new_total"]},
        )
        item.refresh_from_db()
        return Response({
            "data": MenuItemSerializer(item, context={"request": request}).data,
            "meta": result,
        })

    @action(detail=True, methods=["post"], url_path="writeoff_prepared")
    def writeoff_prepared(self, request, pk=None):
        """Phase 8C — списать N испорченных готовых порций batch-блюда.

        POST { qty: int > 0, reason: str } → BatchCookingLog kind=correct (-qty).
        Не списывает сырьё (оно уже было потрачено при заготовке).
        """
        from common.exceptions import BusinessError

        from .services import writeoff_prepared_batch

        item = self.get_object()
        try:
            qty = int(request.data.get("qty") or 0)
        except (TypeError, ValueError):
            raise BusinessError("INVALID_VALUE", "qty должно быть числом", 400)
        reason = (request.data.get("reason") or "").strip()
        result = writeoff_prepared_batch(
            item, qty=qty, reason=reason, user=request.user,
        )
        from apps.audit.services import log_request
        log_request(
            request, "batch_writeoff", target=item,
            payload={"qty": qty, "reason": reason, "new_total": result["new_total"]},
        )
        item.refresh_from_db()
        return Response({
            "data": MenuItemSerializer(item, context={"request": request}).data,
            "meta": result,
        })

    @action(detail=True, methods=["post"], url_path="allow_oversell")
    def allow_oversell(self, request, pk=None):
        """Phase 8D — переключить «продавать в минус».

        Body: {enabled: bool}. При enabled=True снимается авто-стоп для этого
        блюда (даже если остатков нет). При enabled=False — авто-логика
        снова применится при следующем stock movement.
        """
        from apps.audit.services import log_request

        from .services_autostop import reconcile_menu_item_stop

        item = self.get_object()
        enabled = bool(request.data.get("enabled", True))
        item.allow_oversell = enabled
        item.save(update_fields=["allow_oversell", "updated_at"])
        result = reconcile_menu_item_stop(item)
        log_request(
            request, "settings_update", target=item,
            payload={
                "action": "allow_oversell",
                "menu_item": item.name,
                "enabled": enabled,
                "reconcile": result,
            },
        )
        item.refresh_from_db()
        return Response({
            "data": MenuItemSerializer(item, context={"request": request}).data,
            "meta": {"reconcile": result},
        })

    @action(detail=True, methods=["post"], url_path="toggle_tech_card")
    def toggle_tech_card(self, request, pk=None):
        """Phase 8D — переключить auto_consume (учитывать ли техкарту).

        Body: {enabled: bool}. False = блюдо не списывает остатки при close_order
        и не уходит в авто-стоп.
        """
        from apps.audit.services import log_request

        from .services_autostop import reconcile_menu_item_stop

        item = self.get_object()
        enabled = bool(request.data.get("enabled", True))
        item.auto_consume = enabled
        item.save(update_fields=["auto_consume", "updated_at"])
        result = reconcile_menu_item_stop(item)
        log_request(
            request, "settings_update", target=item,
            payload={
                "action": "toggle_tech_card",
                "menu_item": item.name,
                "enabled": enabled,
                "reconcile": result,
            },
        )
        item.refresh_from_db()
        return Response({
            "data": MenuItemSerializer(item, context={"request": request}).data,
            "meta": {"reconcile": result},
        })

    @action(detail=False, methods=["get"], url_path="auto_stopped")
    def auto_stopped_list(self, request):
        """Phase 8D — все блюда, попавшие в авто-стоп из-за нехватки склада."""
        qs = self.get_queryset().filter(auto_stopped=True)
        data = MenuItemSerializer(qs, many=True, context={"request": request}).data
        return Response({"data": data, "meta": {"total": qs.count()}})

    @action(detail=True, methods=["post"], url_path="restore")
    def restore(self, request, pk=None):
        """Вернуть блюдо в продажу: is_available=True + очистка stop-полей."""
        from apps.audit.services import log_request

        item = self.get_object()
        item.is_available = True
        item.stop_reason = ""
        item.stop_until = None
        item.save(
            update_fields=[
                "is_available", "stop_reason", "stop_until", "updated_at",
            ]
        )
        log_request(
            request, "settings_update", target=item,
            payload={"action": "restore", "menu_item": item.name},
        )
        return Response(
            {"data": MenuItemSerializer(item, context={"request": request}).data}
        )


class MenuItemNoteViewSet(
    mixins.ListModelMixin,
    mixins.CreateModelMixin,
    mixins.UpdateModelMixin,
    mixins.DestroyModelMixin,
    viewsets.GenericViewSet,
):
    """CRUD шаблонов комментариев к блюдам.

    Read доступен всем auth (для chip-picker в MenuScreen / Cart).
    Write — только IsCashier (правка через UI Settings).
    """

    serializer_class = MenuItemNoteSerializer
    pagination_class = None
    filterset_fields = ["is_active"]

    def get_queryset(self):
        return MenuItemNote.objects.filter(
            restaurant=self.request.user.restaurant
        )

    def get_permissions(self):
        if self.action in {"create", "update", "partial_update", "destroy"}:
            return [IsCashier()]
        return [IsCashierOrWaiter()]

    def list(self, request, *args, **kwargs):
        qs = self.filter_queryset(self.get_queryset())
        return Response(
            {
                "data": MenuItemNoteSerializer(qs, many=True).data,
                "meta": {"total": qs.count()},
            }
        )

    def perform_create(self, serializer):
        serializer.save(restaurant=self.request.user.restaurant)

    def create(self, request, *args, **kwargs):
        ser = self.get_serializer(data=request.data)
        ser.is_valid(raise_exception=True)
        self.perform_create(ser)
        return Response({"data": ser.data}, status=status.HTTP_201_CREATED)

    def update(self, request, *args, **kwargs):
        partial = kwargs.pop("partial", False)
        instance = self.get_object()
        ser = self.get_serializer(instance, data=request.data, partial=partial)
        ser.is_valid(raise_exception=True)
        ser.save()
        return Response({"data": ser.data})


class ModifierGroupViewSet(
    mixins.ListModelMixin,
    mixins.RetrieveModelMixin,
    mixins.CreateModelMixin,
    mixins.UpdateModelMixin,
    mixins.DestroyModelMixin,
    viewsets.GenericViewSet,
):
    """CRUD групп модификаторов с вложенными опциями.

    Read доступен всем authed (cashier/waiter — для отображения в menu).
    Write — только IsCashier.
    """

    serializer_class = ModifierGroupSerializer
    pagination_class = None
    filterset_fields = ["is_active", "is_required"]

    def get_queryset(self):
        return (
            ModifierGroup.objects
            .filter(restaurant=self.request.user.restaurant)
            .prefetch_related("modifiers")
        )

    def get_permissions(self):
        if self.action in {"create", "update", "partial_update", "destroy"}:
            return [IsCashier()]
        return [IsCashierOrWaiter()]

    def list(self, request, *args, **kwargs):
        qs = self.filter_queryset(self.get_queryset())
        return Response(
            {
                "data": ModifierGroupSerializer(qs, many=True).data,
                "meta": {"total": qs.count()},
            }
        )

    def retrieve(self, request, *args, **kwargs):
        obj = self.get_object()
        return Response({"data": ModifierGroupSerializer(obj).data})

    def perform_create(self, serializer):
        serializer.save(restaurant=self.request.user.restaurant)

    def create(self, request, *args, **kwargs):
        ser = self.get_serializer(data=request.data)
        ser.is_valid(raise_exception=True)
        self.perform_create(ser)
        return Response({"data": ser.data}, status=status.HTTP_201_CREATED)

    def update(self, request, *args, **kwargs):
        partial = kwargs.pop("partial", False)
        instance = self.get_object()
        ser = self.get_serializer(instance, data=request.data, partial=partial)
        ser.is_valid(raise_exception=True)
        ser.save()
        return Response({"data": ser.data})
