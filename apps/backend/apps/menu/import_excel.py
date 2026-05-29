"""Импорт меню из Excel (.xlsx).

Формат файла (sheet «Menu» обязателен):

| Категория  | Блюдо      | Цена  | Эмодзи | Доступно |
|------------|------------|-------|--------|----------|
| Горячее    | Плов       | 45.00 | 🍚     | да       |
| Напитки    | Чай        |  8.00 | 🍵     | да       |

- Строка 1 — заголовки (на русском, регистр-нечувствительно)
- Категории создаются автоматически, если нет
- Цена — Decimal, обязательна
- Эмодзи / Доступно — опциональны
- Существующие блюда (по name + category) обновляются (idempotent)

Использование:
    from apps.menu.import_excel import import_menu_xlsx
    summary = import_menu_xlsx(xlsx_bytes, restaurant=restaurant, user=user)
    # summary = {"created": 5, "updated": 2, "errors": [...]}
"""
from __future__ import annotations

import io
from decimal import Decimal, InvalidOperation
from typing import Any

from django.db import transaction


COLUMN_ALIASES = {
    "category": ("категория", "категории", "category"),
    "name": ("блюдо", "название", "позиция", "name"),
    "price": ("цена", "стоимость", "price"),
    "emoji": ("эмодзи", "иконка", "emoji"),
    "available": ("доступно", "available", "active", "активно"),
}


def _normalize_header(value: Any) -> str:
    return str(value or "").strip().lower()


def _build_header_map(headers: list[str]) -> dict[str, int]:
    """Map of normalized name → column index."""
    out: dict[str, int] = {}
    for i, h in enumerate(headers):
        norm = _normalize_header(h)
        for canonical, aliases in COLUMN_ALIASES.items():
            if norm in aliases:
                out[canonical] = i
                break
    return out


def _parse_decimal(value) -> Decimal | None:
    if value is None or value == "":
        return None
    try:
        return Decimal(str(value).replace(",", ".").strip())
    except (InvalidOperation, ValueError):
        return None


def _parse_bool(value, default: bool = True) -> bool:
    if value is None or value == "":
        return default
    s = str(value).strip().lower()
    if s in ("да", "yes", "y", "true", "1", "✓", "+"):
        return True
    if s in ("нет", "no", "n", "false", "0", "-", "x"):
        return False
    return default


@transaction.atomic
def import_menu_xlsx(
    file_bytes: bytes,
    *,
    restaurant,
    user=None,
    sheet_name: str = "Menu",
) -> dict:
    """Парсит XLSX и upsert'ит категории/блюда.

    Возвращает summary: {created, updated, errors: [{row, message}, ...]}
    """
    try:
        import openpyxl
    except ImportError as exc:
        raise RuntimeError(
            "Для импорта Excel установите openpyxl (`pip install openpyxl`)"
        ) from exc

    from apps.menu.models import Category, MenuItem

    summary = {"created": 0, "updated": 0, "errors": []}

    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        # Берём первый sheet если 'Menu' не найден
        ws = wb.worksheets[0]

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return summary

    headers = list(rows[0])
    col_map = _build_header_map(headers)

    if "category" not in col_map or "name" not in col_map or "price" not in col_map:
        raise ValueError(
            "В файле должны быть колонки: Категория, Блюдо, Цена. "
            f"Найдены: {[_normalize_header(h) for h in headers]}"
        )

    # Кэш категорий для одного импорта (минимизируем запросы)
    cat_cache: dict[str, Category] = {}

    for idx, row in enumerate(rows[1:], start=2):  # row 1 — header
        try:
            cat_name = str(row[col_map["category"]] or "").strip()
            item_name = str(row[col_map["name"]] or "").strip()
            price = _parse_decimal(row[col_map["price"]])
            emoji = str(row[col_map["emoji"]] or "").strip() if "emoji" in col_map else ""
            available = (
                _parse_bool(row[col_map["available"]])
                if "available" in col_map
                else True
            )

            if not cat_name or not item_name or price is None:
                summary["errors"].append({
                    "row": idx,
                    "message": "Пустая категория, название или цена",
                })
                continue

            # Категория
            cat = cat_cache.get(cat_name)
            if cat is None:
                cat, _ = Category.objects.get_or_create(
                    restaurant=restaurant, name=cat_name,
                )
                cat_cache[cat_name] = cat

            # Блюдо
            item, created = MenuItem.objects.update_or_create(
                restaurant=restaurant,
                category=cat,
                name=item_name,
                defaults={
                    "price": price,
                    "emoji": emoji[:8],
                    "is_available": available,
                },
            )
            if created:
                summary["created"] += 1
            else:
                summary["updated"] += 1
        except Exception as exc:
            summary["errors"].append({"row": idx, "message": str(exc)})

    if user is not None:
        from apps.audit.services import audit_log
        audit_log(
            user, "settings_update", target=None,
            payload={
                "action": "menu_import_xlsx",
                "created": summary["created"],
                "updated": summary["updated"],
                "errors": len(summary["errors"]),
            },
        )

    return summary
